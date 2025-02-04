from openpyxl import Workbook

workbook = Workbook()
print(workbook.sheetnames)

planilha = workbook.create_sheet('cadastro')
planilha['A1'] = 10
planilha.cell(row=1, column=2).value = 11

print(planilha['A1'].value)
print(planilha['B1'].value)
