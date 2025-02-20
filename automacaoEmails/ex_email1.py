from openpyxl import load_workbook
from smtplib  import SMTP_SSL
from email.mime.text import MIMEText

import json

inscricoes = {}

with open('config.json', 'r') as arq:
    config = json.load(arq)

servidor = config['servidor']

usuario = config['usuario']
senha = config['senha']


arquivo = load_workbook('inscricoes.xlsx')
planilha = arquivo.active
totalLinhas = planilha.max_row

con = SMTP_SSL(servidor[0], servidor[1])
con.login(usuario, senha)

for linha in range(2,totalLinhas+1):
    inscricao = planilha.cell(linha, 1).value
    nome = planilha.cell(linha, 2).value
    email = planilha.cell(linha, 3).value

    print(f'{email}')
    
    assunto = 'confirmção de inscrição'
    corpo_email = f'ola, {nome} de inscricao numero {inscricao} sua inscrição foi confirmada'

    msg = MIMEText(corpo_email, 'plain')
    msg['Subject'] = assunto
    msg ['From'] = usuario

    con.sendmail(usuario, email, msg.as_string())


